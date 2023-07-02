from auxilary import remove_brackets
from bs4 import BeautifulSoup as bs
from config import *
from csv import writer, reader
from json import load, dump
from openpyxl import load_workbook
from os import makedirs, path
from random import randint
from re import search
from requests import Session, post
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from time import sleep
from tqdm import tqdm
from typing import Tuple, List
# Staff is required for eval(event_and_sex) in aggregate_to_csv function.
from staff.staff import ballet_men, ballet_women, extras_men, extras_women


# TODO: Add tests
# TODO: Add logging
# TODO: Add SQL


class Driver:
    def __init__(self):
        self.rehs_common = {'Реп.', 'Сц. фп.', 'Орк. сцен. реп.', 'Ген. реп.'}
        self.skip_part_common = {'Явка на грим', '(сверка)'}
        options = webdriver.FirefoxOptions()
        options.set_preference("dom.webdriver.enabled", False)
        self.__driver = webdriver.Firefox(options=options)

    # Shorten frequently used Webdriver methods for better readability
    def __dfid(self, ID: str):
        return self.__driver.find_element(By.ID, ID)

    def __dfx(self, xpath: str):
        return self.__driver.find_element(By.XPATH, xpath)

    def __dfsx(self, xpath: str):
        return self.__driver.find_elements(By.XPATH, xpath)

    def __wdwid(self, ID: str):
        wait_time = 60
        WebDriverWait(self.__driver, wait_time).until(
            ec.presence_of_element_located((By.ID, ID)))

    def __login_via_webdriver(self):
        """Login via webdriver. Enter username and password"""
        date_len = 8  # DDMMYYYY
        self.__driver.get(URL_LOGIN)
        self.__wdwid('InputL')
        self.__dfid('InputL').send_keys(USERNAME)
        self.__dfid('InputP').send_keys(PASSWORD, Keys.ENTER)
        self.__wdwid('startDate')

        # Clear previous start date and enter a new one
        date = self.__dfid('startDate')
        for _ in range(date_len):
            date.send_keys(Keys.BACKSPACE)
        date.send_keys(START_DATE)

        # Clear previous final date and enter a new one
        date = self.__dfid('finishDate')
        for _ in range(date_len):
            date.send_keys(Keys.BACKSPACE)
        date.send_keys(FINAL_DATE)

    def __get_codes(self, ballet_extras: str, rehs_keywords: set,
                    skip_part: set):
        """Gets unique event codes and writes them to 2 files with
        corresponding names.
        Click dropdown menu -> choose ballet or extras -> click 'Submit'

        Args:
            ballet_extras: 'ballet' or 'extras'
            rehs_keywords: Get event code if the link text is in these keywords
            skip_part: Skip event code if the link text is in these keywords
        """
        perfs_codes, rehs_codes = set(), set()
        menu_code = '3' if ballet_extras == 'ballet' else '6'
        self.__dfx("//div[@class='btn-group bootstrap-select']").click()
        self.__dfx(f"//li[@data-original-index='{menu_code}']").click()
        self.__dfid('submit').click()

        # The url between ballet and extras does not change due to JS code,
        # so use time.sleep(10) instead of WebDriverWait. Usually 10 sec is
        # enough for a month period, but might be incremented if needed.
        sleep(10)
        links = self.__dfsx("//a[contains(@href, '/Home/MoreInfo/')]")
        for link in links:
            code = search(r'\d+', link.get_attribute('href')).group()
            if any(event in link.text for event in skip_part):
                pass
            elif any(event in link.text for event in rehs_keywords):
                rehs_codes.add(code)
            else:
                perfs_codes.add(code)

        for perf_reh in zip((perfs_codes, rehs_codes), ('perfs', 'rehs')):
            with open(f'{CODES_PATH}{ballet_extras}_{perf_reh[1]}', 'w') as f:
                for code in sorted(perf_reh[0]):
                    f.write(code + '\n')

    def get_all_codes(self):
        """Gets all codes for the ballet and extras"""
        if not path.isdir(CODES_PATH):
            makedirs(CODES_PATH)
        self.__login_via_webdriver()

        # Adds specific keywords for the ballet rehs and ballet skip part
        ballet_rehs = self.rehs_common.union({'+балет'})
        ballet_skip_part = self.skip_part_common.union({'Урок балета'})
        self.__get_codes('ballet', ballet_rehs, ballet_skip_part)

        # Adds specific keywords for the extras rehs and extras skip part
        extras_rehs = self.rehs_common.union({'Тех. работы', 'миманс'})
        extras_skip_part = self.skip_part_common.union({'Занятие +миманс'})
        self.__get_codes('extras', extras_rehs, extras_skip_part)

        self.__driver.quit()


class User:
    """
    These args are frequently used in this class, so I decided not to cram
    every function description and bring them away to a class description.
    Every arg takes strictly one of two values.
    Args:
        ballet_extras (str): 'ballet' or 'extras'
        perfs_rehs (str): 'perfs' or 'rehs'
        men_women (str): 'men' or 'women'
    """

    def __init__(self):
        self.bar_format = \
            '{desc}:|{bar}|{percentage:3.0f}% {n_fmt}/{total_fmt} [{elapsed}]'
        self.session = self.login_via_requests()

    @staticmethod
    def login_via_requests() -> Session:
        """Login via requests.

        Returns:
            requests.Session()
        """
        user_data = {'UserName': USERNAME, 'Password': PASSWORD}
        session = Session()
        session.post(URL_LOGIN, data=user_data)
        return session

    def __parse_table(self, event_code: str, ballet_extras: str) -> List:
        """Parses event table with participant names and roles.

        Args:
            event_code: unique event code from any file in ./data/codes/

        Returns:
            List of names and roles.
        """
        sleep(randint(1, 2))
        menu_code = '?a=9' if ballet_extras == 'ballet' else '?a=11'
        style_tag = 'padding-left: 5px; border: 1px solid gray;'
        event_url = f'{URL_EVENT}{event_code}{menu_code}'
        request = post(event_url, cookies=self.session.cookies)
        return bs(request.text, 'lxml').find_all('td', style=style_tag)

    def count_ballet(self, ballet_codes: set, extras_codes: set,
                     perfs_rehs: str) -> Tuple[dict[set], dict[set]]:
        """Parses html page with ballet codes.

        Args:
            ballet_codes: unique perfs or rehs code from ./data/codes/ballet_*
            extras_codes: unique perfs or rehs code from ./data/codes/extras_*

        Returns:
            Tuple of feature and secure ballet participant names with counters.
        """
        feats, secures = {}, {}
        all_codes = ballet_codes.union(extras_codes)

        for event_code in tqdm(all_codes, desc=f'Ballet {perfs_rehs}',
                               bar_format=self.bar_format, dynamic_ncols=True):
            feat, secure = set(), set()
            for ballet_extras in zip(
                    (ballet_codes, extras_codes), ('ballet', 'extras')):
                if event_code in ballet_extras[0]:
                    self.__count_participants(
                        ballet_extras[1], event_code, feat, secure)
            self.__overflow_participants(feat, secure, feats, secures)

        return feats, secures

    def count_extras(self, extras_codes: set, perfs_rehs: str
                     ) -> Tuple[dict[set], dict[set]]:
        """Parses html page with extras codes.

        Args:
            extras_codes: unique perfs or rehs code from ./data/codes/extras_*

        Returns:
           Tuple of feature and secure extras participant names with counters.
        """
        feats, secures = {}, {}

        for event_code in tqdm(extras_codes, desc=f'Extras {perfs_rehs}',
                               bar_format=self.bar_format, dynamic_ncols=True):
            feat, secure = set(), set()
            self.__count_participants('extras', event_code, feat, secure)
            self.__overflow_participants(feat, secure, feats, secures)

        return feats, secures

    def __count_participants(self, ballet_extras: str, event_code: str,
                             feat: set, secure: set):
        """Counts number of participations in each event.

        Args:
            event_code: unique perfs or rehs code from ./data/codes/*
            feat: names of featured participants
            secure: names of secure participants
        """
        table = [person.text.strip() for person in
                 self.__parse_table(event_code, ballet_extras)]
        people = self.__distribute_participants(table)
        feat.update(people[0])
        secure.update(people[1])

    @staticmethod
    def __overflow_participants(feat: set, secure: set, events_feat: dict,
                                events_secure: dict):
        """Overflows featured participants from secure event to feat event and
        add to corresponding dictionaries with event counters.

        Args:
            feat: names of featured participants
            secure: names of secure participants
            events_feat: dictionary with featured participants counter
            events_secure: dictionary with secure participants counter
        """
        secure.difference_update(feat)
        for person in feat:
            events_feat[person] = events_feat.get(person, 0) + 1
        for person in secure:
            events_secure[person] = events_secure.get(person, 0) + 1

    @staticmethod
    def __distribute_participants(table: list) -> Tuple[set[str], set[str]]:
        """Distributes participants in 2 sets: by feature and secure. Skip the
        row, if name is in the skip part. Remove brackets and text inside.

        Args:
            table: list of 3 columns: Role | Feature | Secure.

        Returns:
            Tuple of feature and secure sets with participant names.
        """
        feat, secure = set(), set()
        skip_part = {
            'режиссер', 'режисер', 'миманс', 'педагог', 'инспектор',
            'концермейстер', 'концертмейстер', 'руководитель балетной труппы'
        }

        for i in range(0, len(table), 3):
            if not any(part in table[i].lower() for part in skip_part):
                feat_column = table[i + 1].split(',')
                for person in feat_column:
                    feat.add(remove_brackets(person))
                secure_column = table[i + 2].split(',')
                for person in secure_column:
                    secure.add(remove_brackets(person))

        return feat, secure

    def __get_all_codes(self, ballet_extras: str, perfs_rehs: str
                        ) -> Tuple[dict[set], dict[set]]:
        """Gets all codes. For the ballet we need both ballet and extras codes.
        For the extras we need only extras codes.

        Returns:
            Tuple of feature and secure participant names with counters.
        """
        extras_codes = self.__get_codes('extras', perfs_rehs)
        if ballet_extras == 'ballet':
            ballet_codes = self.__get_codes('ballet', perfs_rehs)
            return self.count_ballet(ballet_codes, extras_codes, perfs_rehs)
        return self.count_extras(extras_codes, perfs_rehs)

    @staticmethod
    def __get_codes(ballet_extras: str, perfs_rehs: str) -> set:
        """Get ballet or extras perfs or rehs codes from corresponding file.

        Returns:
            Set of codes
        """
        with open(f'{CODES_PATH}{ballet_extras}_{perfs_rehs}') as codes:
            return set(codes.read().split())

    def run_parser(self, ballet_extras: str, perfs_rehs: str):
        """Writes ballet and extras feature and secure to json file in format:
        {name: events' counter}"""
        feat, secure = self.__get_all_codes(ballet_extras, perfs_rehs)

        if not path.isdir(JSON_PATH):
            makedirs(JSON_PATH)

        for feat_secure in zip((feat, secure), ('feat', 'secure')):
            # Remove empty values
            if '' in feat_secure[0]:
                del feat_secure[0]['']
            json_file = f'{JSON_PATH}{ballet_extras}_' \
                        f'{feat_secure[1]}_{perfs_rehs}.json'
            with open(json_file, 'w', encoding='utf-8') as file:
                dump(feat_secure[0], file, indent=4, ensure_ascii=False)

    @staticmethod
    def aggregate_to_csv(ballet_extras: str, men_women: str):
        """Aggregates json to csv and writes to file."""

        # Create a new table with the names from a staff module
        event_and_sex = f'{ballet_extras}_{men_women}'
        csv_file = f'{CSV_PATH}{event_and_sex}_{DATES}.csv'
        new_table = {person: [0, 0, 0, 0] for person in eval(event_and_sex)}

        # Column index for each of 4 columns
        column_index = 0
        for perf_reh in ('perfs', 'rehs'):
            for feat_secure in ('feat', 'secure'):
                json_file = f'{JSON_PATH}{ballet_extras}_' \
                            f'{feat_secure}_{perf_reh}.json'
                with open(json_file, encoding='utf-8') as file:
                    table = load(file)
                for person in eval(event_and_sex):
                    if person in table:
                        new_table[person][column_index] += table[person]
                column_index += 1

        # Remove empty values
        for key, value in tuple(new_table.items()):
            if not any(value):
                del new_table[key]

        # Create rows with values ('C' - counter; sorted by perf feat C):
        # Name, perf feat C, perf secure C, reh feat C, reh secure C
        rows = [[key, *value] for key, value in
                sorted(new_table.items(), key=lambda x: x[1], reverse=True)]

        if not path.isdir(CSV_PATH):
            makedirs(CSV_PATH)
        with open(csv_file, 'w', encoding='utf-8', newline='\n') as file:
            writer(file).writerows(rows)

    def write_to_xls(self, ballet_extras: str, men_women: str):
        """Writes csv to prepared xls template."""

        xls_path = f'./data/{DATES}/xls/'
        xls_file = f'{xls_path}{ballet_extras}_{men_women}_{DATES}.xlsx'
        csv_file = f'{CSV_PATH}{ballet_extras}_{men_women}_{DATES}.csv'
        workbook = load_workbook('./templates/excel/template.xlsx')
        worksheet = workbook.active
        worksheet.cell(1, 1).value = DATES
        last_row_index = 66  # extras_men

        with open(csv_file, encoding='utf-8', newline='\n') as file:
            for row_index, row in enumerate(reader(file), start=11):
                for col_index, value in enumerate(row, start=1):
                    worksheet.cell(row_index, col_index).value = value
            # Remove excessive values
            for row in range(row_index, last_row_index):
                if worksheet.cell(row, 1).value is None:
                    worksheet.cell(row, 6).value = None

        if not path.isdir(xls_path):
            makedirs(xls_path)
        workbook.save(filename=xls_file)


def main():
    # Login via webdriver and collect all codes
    # Driver().get_all_codes()

    user = User()
    # Parse through events to get json
    for ballet_extras in ('ballet', 'extras'):
        for perfs_rehs in ('perfs', 'rehs'):
            user.run_parser(ballet_extras, perfs_rehs)

    # Aggregate and convert json to csv. Write csv to xls
    for ballet_extras in ('ballet', 'extras'):
        for men_women in ('men', 'women'):
            user.aggregate_to_csv(ballet_extras, men_women)
            user.write_to_xls(ballet_extras, men_women)


if __name__ == '__main__':
    main()
