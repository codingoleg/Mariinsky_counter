"""
These args are frequently used in this module, so I decided not to cram
every function description and bring them away to a class description.
Every arg takes strictly one of two values:
    event_type (str): 'ballet' or 'extras'
    action_type (str): 'perfs' or 'rehs'
    participation_type (str): 'feat' or 'secure'
    gender (str): 'men' or 'women'
"""

import csv
import json
import logging
import os
import random
import re
import time
from typing import Tuple, List, Dict, Set

import bs4
import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from tqdm import tqdm

import xpaths as xp
from constants import DATES, CODES_PATH, JSON_PATH, CSV_PATH, XLS_PATH, WORKBOOK_PATH
from constants import BALLET, EXTRAS, FEAT, SECURE, PERFS, REHS, BAR_FORMAT
from config import USERNAME, PASSWORD, URL_LOGIN, URL_EVENT, START_DATE, FINAL_DATE
from utils import remove_brackets
# staff.staff is required for eval(event_and_sex) in aggregate_to_csv function.
from staff.staff import ballet_men, ballet_women, extras_men, extras_women

logging.basicConfig(
    format='%(asctime)s - %(message)s',
    datefmt='%d-%b-%y %H:%M:%S',
    level=logging.INFO
)


class Driver:
    def __init__(self):
        self.rehs_common = {'Реп.', 'Сц. фп.', 'Орк. сцен. реп.', 'Ген. реп.'}
        self.skip_part_common = {'Явка на грим', '(сверка)'}
        self.__driver: webdriver

    # Shorten frequently used Webdriver methods for better readability
    def __dfid(self, ID: str):
        return self.__driver.find_element(By.ID, ID)

    def __dfx(self, xpath: str):
        return self.__driver.find_element(By.XPATH, xpath)

    def __dfsx(self, xpath: str):
        return self.__driver.find_elements(By.XPATH, xpath)

    def __wdwid(self, ID: str, wait_time: int = 60):
        WebDriverWait(self.__driver, wait_time).until(ec.presence_of_element_located((By.ID, ID)))

    def __login(self) -> None:
        """
        Runs webdriver. Enters username and password. On the next page enters
        start and final DATES.
        """
        date_len = 8  # DDMMYYYY
        options = webdriver.FirefoxOptions()
        options.set_preference("dom.webdriver.enabled", False)
        self.__driver = webdriver.Firefox(options=options)
        logging.info('Webdriver starts')
        self.__driver.get(URL_LOGIN)
        self.__wdwid(xp.username)
        self.__dfid(xp.username).send_keys(USERNAME)
        logging.info('Enters username')
        self.__dfid(xp.password).send_keys(PASSWORD, Keys.ENTER)
        logging.info('Enters password')
        self.__wdwid(xp.start_date)

        # Clear previous start date and enter a new one
        date = self.__dfid(xp.start_date)
        for _ in range(date_len):
            date.send_keys(Keys.BACKSPACE)
        date.send_keys(START_DATE)
        logging.info('Enters start date')

        # Clear previous final date and enter a new one
        date = self.__dfid(xp.final_date)
        for _ in range(date_len):
            date.send_keys(Keys.BACKSPACE)
        date.send_keys(FINAL_DATE)
        logging.info('Enters final date')

    def __get_codes(self, event_type: str, rehs_keywords: Set, skip_part: Set) -> None:
        """
        Gets unique event codes and writes them to 2 files with
        corresponding names.
        Click dropdown menu -> choose ballet or extras -> click 'Submit'
        Args:
            rehs_keywords: Get event code if the link text is in these keywords
            skip_part: Skip event code if the link text is in these keywords
        """
        perfs_codes, rehs_codes = set(), set()
        menu_code = '3' if event_type == BALLET else '6'
        self.__dfx(xp.dropdown_menu).click()
        self.__dfx(f"//li[@data-original-index='{menu_code}']").click()
        self.__dfid(xp.submit).click()
        logging.info(f'Choose {event_type} from menu')

        # The url between ballet and extras does not change due to JS code,
        # so use time.sleep(10) instead of WebDriverWait. Usually 10 sec is
        # enough for a month period, but might be incremented if needed.
        time.sleep(10)
        links = self.__dfsx(xp.links)
        for link in links:
            code = re.search(r'\d+', link.get_attribute('href')).group()
            if any(event in link.text for event in skip_part):
                pass
            elif any(event in link.text for event in rehs_keywords):
                rehs_codes.add(code)
            else:
                perfs_codes.add(code)
        logging.info(f'{event_type} links received')

        for action_type in zip((perfs_codes, rehs_codes), (PERFS, REHS)):
            with open(f'{CODES_PATH}{event_type}_{action_type[1]}', 'w') as file:
                for code in sorted(action_type[0]):
                    file.write(code + '\n')

    def get_all_codes(self) -> None:
        """Gets all codes for the ballet and extras"""
        if not os.path.isdir(CODES_PATH):
            os.makedirs(CODES_PATH)

        self.__login()

        # Adds specific keywords for the ballet rehs and ballet skip part
        ballet_rehs = self.rehs_common.union({'+балет'})
        ballet_skip_part = self.skip_part_common.union({'Урок балета'})
        self.__get_codes(BALLET, ballet_rehs, ballet_skip_part)
        logging.info(f'{BALLET} codes received')

        # Adds specific keywords for the extras rehs and extras skip part
        extras_rehs = self.rehs_common.union({'Тех. работы', 'миманс'})
        extras_skip_part = self.skip_part_common.union({'Занятие +миманс'})
        self.__get_codes(EXTRAS, extras_rehs, extras_skip_part)
        logging.info(f'{EXTRAS} codes received')

        self.__driver.quit()
        logging.info('Webdriver quits')


class User:
    def __init__(self):
        self.session = self.login()

    @staticmethod
    def login() -> requests.Session:
        """
        Login via requests.
        Returns:
            requests.Session()
        """
        session = requests.Session()
        session.post(URL_LOGIN, {'UserName': USERNAME, 'Password': PASSWORD})
        logging.info('Session received')

        return session

    def __parse_table(self, event_code: str, event_type: str) -> List:
        """
        Parses event table with participant names and roles.
        Args:
            event_code: unique event code from any file in ./data/codes/
        Returns:
            List of names and roles.
        """
        time.sleep(random.randint(1, 2))
        menu_code = '?a=9' if event_type == BALLET else '?a=11'
        style_tag = 'padding-left: 5px; border: 1px solid gray;'
        event_url = URL_EVENT + event_code + menu_code
        request = requests.post(event_url, cookies=self.session.cookies)

        return bs4.BeautifulSoup(request.text, 'lxml').find_all('td', style=style_tag)

    def count_ballet(self, ballet_codes: Set, extras_codes: Set, action_type: str
                     ) -> Tuple[dict[str:int], dict[str:int]]:
        """
        Parses html page with ballet codes.
        Args:
            ballet_codes: unique perfs or rehs code from ./data/codes/ballet_*
            extras_codes: unique perfs or rehs code from ./data/codes/extras_*
        Returns:
            Tuple of feature and secure ballet participant names with counters.
        """
        feats, secures = {}, {}
        all_codes = ballet_codes.union(extras_codes)

        for event_code in tqdm(all_codes, desc=f'{BALLET} {action_type}', bar_format=BAR_FORMAT, dynamic_ncols=True):
            feat, secure = set(), set()
            for event_type in zip((ballet_codes, extras_codes), (BALLET, EXTRAS)):
                if event_code in event_type[0]:
                    self.__count_participants(event_type[1], event_code, feat, secure)
            self.__overflow_participants(feat, secure, feats, secures)

        return feats, secures

    def count_extras(self, extras_codes: Set, action_type: str) -> Tuple[dict[str:int], dict[str:int]]:
        """
        Parses html page with extras codes.
        Args:
            extras_codes: unique perfs or rehs code from ./data/codes/extras_*
        Returns:
            Tuple of feature and secure extras participant names with counters.
        """
        feats, secures = {}, {}

        for event_code in tqdm(extras_codes, desc=f'{EXTRAS} {action_type}', bar_format=BAR_FORMAT, dynamic_ncols=True):
            feat, secure = set(), set()
            self.__count_participants(EXTRAS, event_code, feat, secure)
            self.__overflow_participants(feat, secure, feats, secures)

        return feats, secures

    def __count_participants(self, event_type: str, event_code: str, feat: Set, secure: Set) -> None:
        """
        Counts number of participations in each event.
        Args:
            event_code: unique perfs or rehs code from ./data/codes/*
            feat: names of featured participants
            secure: names of secure participants
        """
        table = [person.text.strip() for person in self.__parse_table(event_code, event_type)]
        person_feat, person_secure = self.__distribute_participants(table)
        feat.update(person_feat)
        secure.update(person_secure)

    @staticmethod
    def __overflow_participants(feat: Set, secure: Set, events_feat: Dict, events_secure: Dict) -> None:
        """
        Overflows featured participants from secure event to feat event and
        add to corresponding dictionaries with event counters.
        Args:
            feat: names of featured participants
            secure: names of secure participants
            events_feat: dictionary with featured participants counter
            events_secure: dictionary with secure participants counter
        """
        secure.difference_update(feat)
        for person in feat:
            events_feat[person] = events_feat.get(person, 0) + 1
        for person in secure:
            events_secure[person] = events_secure.get(person, 0) + 1

    @staticmethod
    def __distribute_participants(table: List) -> Tuple[Set[str], Set[str]]:
        """
        Distributes participants in 2 sets: by feature and secure. Skip the
        row, if name is in the skip part. Remove brackets and text inside.
        Args:
            table: list of 3 columns: Role | Feature | Secure.
        Returns:
            Tuple of feature and secure sets with participant names.
        """
        feat, secure = set(), set()
        skip_part = {
            'режиссер', 'режисер', 'миманс', 'педагог', 'инспектор',
            'концермейстер', 'концертмейстер', 'руководитель балетной труппы'
        }

        for i in range(0, len(table), 3):
            if not any(part in table[i].lower() for part in skip_part):

                feat_column = table[i + 1].split(',')
                for person in feat_column:
                    feat.add(remove_brackets(person))

                secure_column = table[i + 2].split(',')
                for person in secure_column:
                    secure.add(remove_brackets(person))

        return feat, secure

    def __get_all_codes(self, event_type: str, action_type: str) -> Tuple[dict[str:int], dict[str:int]]:
        """
        Gets all codes. For the ballet we need both ballet and extras codes.
        For the extras we need only extras codes.
        Returns:
            Tuple of feature and secure participant names with counters.
        """
        with open(f'{CODES_PATH}{EXTRAS}_{action_type}') as codes:
            extras_codes = set(codes.read().split())

        if event_type == BALLET:
            with open(f'{CODES_PATH}{BALLET}_{action_type}') as codes:
                ballet_codes = set(codes.read().split())

            return self.count_ballet(ballet_codes, extras_codes, action_type)

        return self.count_extras(extras_codes, action_type)

    def run_parser(self, event_type: str, action_type: str) -> None:
        """
        Writes ballet and extras feature and secure to json file in format:
        {name: events' counter}
        """
        feat, secure = self.__get_all_codes(event_type, action_type)

        if not os.path.isdir(JSON_PATH):
            os.makedirs(JSON_PATH)

        for participation_type in zip((feat, secure), (FEAT, SECURE)):
            # Remove empty values
            if '' in participation_type[0]:
                del participation_type[0]['']
            json_file = f'{JSON_PATH}{event_type}_{participation_type[1]}_{action_type}.json'
            with open(json_file, 'w', encoding='utf-8') as file:
                json.dump(participation_type[0], file, indent=4, ensure_ascii=False)


class Data:
    @staticmethod
    def aggregate_to_csv(event_type: str, gender: str) -> None:
        """Aggregates json to csv and writes to file."""

        # Create a new table with the names from a staff module
        event_and_sex = f'{event_type}_{gender}'
        csv_file = f'{CSV_PATH}{event_and_sex}_{DATES}.csv'
        new_table = {person: [0, 0, 0, 0] for person in eval(event_and_sex)}

        # Column index for each of 4 columns
        column_index = 0
        for action_type in (PERFS, REHS):
            for participation_type in (FEAT, SECURE):
                json_file = f'{JSON_PATH}{event_type}_{participation_type}_{action_type}.json'
                with open(json_file, encoding='utf-8') as file:
                    table = json.load(file)
                for person in eval(event_and_sex):
                    if person in table:
                        new_table[person][column_index] += table[person]
                column_index += 1

        # Remove empty values
        for key, value in tuple(new_table.items()):
            if not any(value):
                del new_table[key]

        # Create rows with values ('C' - counter; sorted by perf feat C):
        # Name, perf feat C, perf secure C, reh feat C, reh secure C
        rows = [[key, *value] for key, value in sorted(new_table.items(), key=lambda x: x[1], reverse=True)]

        if not os.path.isdir(CSV_PATH):
            os.makedirs(CSV_PATH)
        with open(csv_file, 'w', encoding='utf-8', newline='\n') as file:
            csv.writer(file).writerows(rows)

    @staticmethod
    def write_to_xls(event_type: str, gender: str) -> None:
        """Writes csv to prepared xls template."""
        xls_file = f'{XLS_PATH}{event_type}_{gender}_{DATES}.xlsx'
        csv_file = f'{CSV_PATH}{event_type}_{gender}_{DATES}.csv'
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        worksheet = workbook.active
        worksheet.cell(1, 1).value = DATES
        # Index of the longest row in all the documents
        last_row_index = 66

        with open(csv_file, encoding='utf-8', newline='\n') as file:

            for row_index, row in enumerate(csv.reader(file), start=11):
                for col_index, value in enumerate(row, start=1):
                    worksheet.cell(row_index, col_index).value = value

            # Remove excessive values
            for row in range(row_index, last_row_index):
                if worksheet.cell(row, 1).value is None:
                    worksheet.cell(row, 6).value = None

        if not os.path.isdir(XLS_PATH):
            os.makedirs(XLS_PATH)
        workbook.save(filename=xls_file)
